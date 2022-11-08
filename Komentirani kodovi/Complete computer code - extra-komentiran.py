#Moduli za serial communication (komunikacije između računala i microbita) + datum i vrijeme
import serial
import datetime
import time
from serial.tools.list_ports import comports

#Moduli za Excel
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import json

#Ovo je potrebno za slanje mailova
import smtplib
from email.message import EmailMessage

print("Priključite microbit prije nego što započnete") #Traži da se microbit baza spoji na računalo
izbor = input("Želite li primati obavjesti o povišenim temperaturama? DA ili NE: ") 

#Odluka korisnika vezano uz primanje notifikacija
while izbor != "DA" and izbor != "NE":
    izbor = input("Pogreška u unosu, napišite ili DA ili NE: ")

if izbor == "DA":
    mail = input("Upišite e-mail adresu kako biste dobivali obavjesti: ")
    limit = input("Upišite temperaturu u °C iznad koje biste htjeli dobivati notifikacije: ")
else:
    limit = "100"
    

#Početak programa
ulaz = input("Napišite START kako biste započeli: ")

while True:
    if ulaz == "START":

        #Provjerava se je li microbit spojen na usb port i na koji
        for p in comports():
            try:
                ser = serial.Serial(p.name, 9600, timeout=3)
                break
            except:
                if comports().index(p) == len(comports())-1:
                    print("Mikrobit nije spojen ni na jedan aktivan port. Ponovo pokrenite program i pokušajte ponovo.")
                    quit()
                continue
        
        #Računalo šalje bazi vrijeme i datum
        time.sleep(2)
        slanje = (f"{datetime.datetime.now().hour} {datetime.datetime.now().minute} {datetime.datetime.now().second} {datetime.datetime.now().day} {datetime.datetime.now().month} {datetime.datetime.now().year} {limit}")
        ser.write(slanje.encode())

        #Čeka se odgovor microbita (baze)
        while True:
            data = ser.readline()
            encoding = "utf-8"
            data = str(data, encoding, errors="ignore")
            print(data)
            if data:
                #Dobiveni string se prebacuje u listu
                data = data.split("|")
                
                #"Gatherd data" string se uklanja iz liste, prepoznaje se jesu li to završni podatci
                if data[-1] == "Gathered data":
                    data.pop(-1)
                    
                    #Daljnje mijenjanje oblika podataka
                    data2 = []
                    for element in data:
                        element = element.split(",")
                        data2.append(element)
                
                    data = data2.copy()
                
                    #Stvaranje rječnika ovog oblika: {id:{date:{time:temp}}}
                    newData = {}
                    newData2 = {}
                    newData3 = {}
                    lista = []
                    rječnik = {}
                    rješeno = False

                    #Određivanje maksimalnog broja aktivnih microbita
                    for x in range(len(data)):
                        lista.append(int(data[x][0]))

                    n = max(lista)

                    lista = []

                    #Izlučivanje "id" microbita i povezivanje podatak po "id" microbita
                    for i in range(1, n+1):
                        lista = []
                        for j in data:
                            if int(j[0]) == i:
                                lista.append(j[1:])
                        newData.update({str(i):lista})

                    #Dobivanje završnog oblika rječnika
                    for broj, ostalo in newData.items():
                        datumi = set()
                        vrijemeTemp = {}
                        newData2 = {}
                        for član in ostalo:
                            datumi.add(član[0])

                        for datum in datumi:
                            for x in ostalo:
                                if x[0] == datum:
                                    vrijemeTemp.update({x[1]:x[2]})
                            newData2.update({datum:vrijemeTemp})
                        newData3.update({broj:newData2})

                    data = newData3.copy()
                    print(data)
                    
                    #Prijenos podataka u excel
                    
                    # Priprema za stvaranje Excel tablice, sprema podatke i stvara Excel tablicu
                    wb = Workbook() 
                    ws = wb.active
                    ws.title = "Temperature"

                    br_ap = 1 #Broj zauzetih redova

                    temperature = []
                    vremena = []
                   
                     #Razbija bivši rječnik i djeli određene podatke u svoje posebne liste
                    for microbit, dan in data.items():
                        for dan, vrijeme in data[microbit].items():
                            for vrijeme, temperatura in data[microbit][dan].items():
                                temperature.append(temperatura)
                                vremena.append(vrijeme)

                    M = int(microbit) #Broj sveukupnih microbita

                    #Uklanja duple zapise vremena (bolje objašnjeno pod "funkcije-komentirane.py")
                    def my_function(x):
                      return list(dict.fromkeys(x))
                    vremena = my_function(vremena)
                    
                    
                    n = len(vremena) #Finalni broj vremena.
                    br = 0 
                    brp = 1
                    m = (len(temperature)//M)
                    p = m
                    lista = []
                    
                    #Iz liste "temperature" izvlače se pojedinačne temperature koje se kasnije kronološkim redoslijedom zapisuju u Excel
                    for microbit in data:
                        for i in range(0, n):
                            vrijeme = vremena[i]
                            temp = temperature[br:m:n]
                            temp = [eval(i) for i in temp]
                            if brp % n != 0:
                                br += 1
                            if brp % n == 0:
                                br = m
                                m = m + p
                            brp += 1
                            temp.insert(0, vrijeme)
                            print(temp)
                            lista.append(temp)

                    l = 0 #Index na kojem će se temperature izvlačiti iz liste "Lista"
                    
                    #Zapisuje potrebne podatke u Excel (id, datume, vremena i temperature)
                    for microbit in data:
                        mic = ['MICROBIT_' + microbit]
                        ws.append([''] + mic)

                        datumi = [''] + list(data[microbit].keys())
                        ws.append(datumi)

                        br_ap2 = br_ap + 1
                        ws.merge_cells(f"{'A' + str(br_ap)}:{'A' + str(br_ap2)}")

                        for i in range(0, n):
                            ws.append(lista[l])
                            l += 1
                            br_ap += 1

                        ws.append([''])
                        ws.append([''])
                        br_ap += 4


                    wb.save("NovTemp.xlsx") #Sprema dokument

                #U slučaju prelaska temperature šalje upozorenje
                elif len(data) == 3:
                    if data[-2] == "Warning":
                        if izbor == "NE":
                            data = ""
                            continue
                        else:
                            data[0] = data[0].split(",")
                            poruka = (f"High temperature warning - id:{data[0][0]}, date:{data[0][1]}, time:{data[0][2]}, temp:{data[0][3]}")

                            #Stvaranje e-mail poruke
                            msg = EmailMessage()
                            msg["From"] = "microbit.termostat@outlook.com"
                            msg["To"] = mail
                            msg["Subject"] = "Temperature warning"
                            msg.set_content(poruka)

                            #Slanje e-mail poruke
                            s = smtplib.SMTP("smtp-mail.outlook.com", port=587)
                            s.starttls()
                            s.login("microbit.termostat@outlook.com", "Microbit314")
                            s.sendmail("microbit.termostat@outlook.com" , mail, msg.as_string())
                            s.quit()
                            
    else:
        #Prevencija u slučaju da korisnik ne upiše pravilno potrebnu rijec ("START")
        while ulaz != "START":
            ulaz = input("Pogreška u unosu, napišite START kako biste započeli")
